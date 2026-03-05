"""Industrial Guided Task (IGT) upload pipeline — ServiceNow ICW module.

Reads source documents and creates IGT Standard records in ServiceNow:
  table:     sn_icw_igt_standard    (extends sn_icw_std_standard)
  questions: sn_smart_asmt_question (one Checkbox question per procedure step,
             grouped in one sn_smart_asmt_section per document)

Each procedure step from the source document becomes an assessment question
that the operator checks off when they complete that step in the field.

Supported source files
  .docx   — procedure table rows extracted as steps
             (detects ETAPE | METHODE | PHOTO column layout automatically)
             PHOTO column images are uploaded as question guidance attachments
  .xlsx   — header row auto-detected; each data row becomes one IGT step
             expected columns: Step/Étape | Instructions/Méthode | Notes

Triggered automatically by upload_all.py when:
  - the source folder name contains "igt" (case-insensitive), OR
  - the --igt flag is passed on the command line

Environment variables (same .env as KB pipelines, plus):
  SN_ICW_ASSIGNMENT_TYPE  — cmdb_assignment_type value for new standards
                            (default: "equipment")
  SN_IGT_QUESTION_TYPE    — sys_id of sn_smart_asmt_question_type to use
                            (default: Checkbox fb759e8b7771211058119a372e5a99b3)
"""

import os
import sys
import re
import requests
from pathlib import Path
from dotenv import load_dotenv

_script_dir = Path(__file__).parent
load_dotenv(_script_dir.parent / ".env")           # shared: credentials + KB config
load_dotenv(_script_dir / ".env", override=True)   # IGT-specific: references (takes priority)

# Allow imports from sibling sn-kb-upload and parent
sys.path.insert(0, str(_script_dir.parent))
sys.path.insert(0, str(_script_dir.parent / "sn-kb-upload"))

from sn_kb_shared import (
    upload_attachment,
    create_igt_standard,
    get_igt_assessment_template,
    create_igt_section,
    create_igt_question,
    update_igt_question,
    create_igt_response_option,
)


# ── Constants ─────────────────────────────────────────────────────────────────

SUPPORTED_EXTENSIONS = {".docx", ".xlsx"}

# Column keywords used when auto-detecting procedure table headers
_ETAPE_KW    = ("etape", "étape", "step", "tâche", "task", "operation", "opération")
_METHODE_KW  = ("method", "méthode", "instruction", "description", "procédure", "procedure")
_PHOTO_KW    = ("photo", "image", "picture", "illustration", "figure", "img")


# ── Config ────────────────────────────────────────────────────────────────────

def get_config():
    """Read and validate required environment variables."""
    instance      = os.environ.get("SN_INSTANCE", "").rstrip("/")
    username      = os.environ.get("SN_USERNAME", "")
    password      = os.environ.get("SN_PASSWORD", "")
    assign_type   = os.environ.get("SN_ICW_ASSIGNMENT_TYPE", "equipment")

    missing = []
    if not instance: missing.append("SN_INSTANCE")
    if not username: missing.append("SN_USERNAME")
    if not password: missing.append("SN_PASSWORD")

    if missing:
        print(f"ERROR: Missing environment variables: {', '.join(missing)}")
        sys.exit(1)

    return instance, username, password, assign_type


# ── Step extraction ────────────────────────────────────────────────────────────

def _col_index(headers, *keywords):
    """Return the first column index whose lowercased header contains any keyword."""
    for i, h in enumerate(headers):
        hl = h.strip().lower()
        if any(kw in hl for kw in keywords):
            return i
    return -1


def _extract_cell_images(cell, doc):
    """Return a list of (bytes, content_type) for all images embedded in a table cell."""
    try:
        from docx.oxml.ns import qn
    except ImportError:
        return []
    images = []
    for blip in cell._element.iter(qn("a:blip")):
        r_embed = blip.get(qn("r:embed"))
        if r_embed:
            try:
                part = doc.part.related_parts[r_embed]
                images.append((part.blob, part.content_type))
            except (KeyError, AttributeError):
                pass
    return images


def extract_steps_from_docx(file_path):
    """Parse ETAPE | METHODE | PHOTO procedure tables from a DOCX file.

    Returns a list of dicts: {title, instructions, order, photos}.
      title        — ETAPE column (step category / context label)
      instructions — METHODE column (the actual instruction text = question label)
      order        — row sequence number (1-based)
      photos       — list of (bytes, content_type) from the PHOTO column

    Handles merged/rowspan ETAPE cells by carrying the last non-empty title forward.
    Skips tables that do not have recognisable step/method columns.
    """
    try:
        from docx import Document
    except ImportError:
        print("  Warning: python-docx not installed — step extraction skipped")
        return []

    try:
        doc = Document(file_path)
    except Exception as e:
        print(f"  Warning: could not open DOCX for step extraction: {e}")
        return []

    steps = []
    order = 0
    seen = set()   # deduplicate rows shared across continuation tables

    for table in doc.tables:
        if not table.rows:
            continue

        # Auto-detect column positions from the first row
        header_texts = [cell.text.strip() for cell in table.rows[0].cells]
        etape_col   = _col_index(header_texts, *_ETAPE_KW)
        methode_col = _col_index(header_texts, *_METHODE_KW)
        photo_col   = _col_index(header_texts, *_PHOTO_KW)

        if etape_col == -1 and methode_col == -1:
            continue  # not a procedure table

        prev_etape = ""
        for row in table.rows[1:]:
            cells = row.cells
            etape   = cells[etape_col].text.strip()   if 0 <= etape_col   < len(cells) else ""
            methode = cells[methode_col].text.strip() if 0 <= methode_col < len(cells) else ""

            # Carry forward ETAPE when the cell is merged (rowspan)
            if not etape:
                etape = prev_etape
            else:
                prev_etape = etape

            if not (etape or methode):
                continue

            key = (etape[:80], methode[:80])
            if key in seen:
                continue
            seen.add(key)

            # Extract images from PHOTO column if present
            photos = []
            if 0 <= photo_col < len(cells):
                photos = _extract_cell_images(cells[photo_col], doc)

            order += 1
            steps.append({
                "title":        etape[:255],
                "instructions": methode,
                "order":        order,
                "photos":       photos,
            })

    return steps


def extract_steps_from_excel(file_path):
    """Read rows from the first sheet of an Excel file as IGT steps.

    The header row is auto-detected. Expected column names (case-insensitive):
      Step title    : Step | Étape | Etape | Name | Nom | Task
      Instructions  : Instructions | Method | Méthode | Description
      Notes (opt.)  : Notes | Remarks | Remarques | Comments

    Falls back to column A = title, column B = instructions when no
    matching headers are found.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  Warning: openpyxl not installed — Excel step extraction skipped")
        return []

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  Warning: could not open Excel for step extraction: {e}")
        return []

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return []

    # Auto-detect header columns
    headers = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    title_col  = _col_index(headers, "step", "étape", "etape", "name", "nom", "task", "tâche")
    instr_col  = _col_index(headers, "instruction", "method", "méthode", "description", "procédure")
    notes_col  = _col_index(headers, "note", "remark", "remarque", "comment")

    if title_col == -1 and instr_col == -1:
        title_col, instr_col = 0, 1  # fallback: A=title, B=instructions

    def _cell(row, idx):
        if idx < 0 or idx >= len(row) or row[idx] is None:
            return ""
        return str(row[idx]).strip()

    steps = []
    for order, row in enumerate(rows[1:], start=1):
        title  = _cell(row, title_col)
        instr  = _cell(row, instr_col)
        notes  = _cell(row, notes_col)

        if notes:
            instr = f"{instr}\n\nNote: {notes}".strip()

        if not (title or instr):
            continue

        steps.append({
            "title":        title[:255],
            "instructions": instr,
            "order":        order,
        })

    return steps


# ── Main processing ────────────────────────────────────────────────────────────

def process_igt(file_path, instance, username, password, assign_type):
    """Process one source file → create IGT Standard + questions in ServiceNow.

    Each procedure step becomes one Checkbox question:
      label             = METHODE instructions (the actual task the operator reads)
      guidance_statement= photo(s) from the PHOTO column, uploaded as attachments

    Returns (igt_number, error_message).  error_message is None on success.
    """
    ext   = file_path.suffix.lower()
    title = file_path.stem.replace("_", " ").replace("-", " ").title()

    # ── 1. Extract procedure steps ─────────────────────────────────────────────
    steps = []
    if ext == ".docx":
        steps = extract_steps_from_docx(file_path)
        if steps:
            photo_count = sum(len(s["photos"]) for s in steps)
            photo_info  = f", {photo_count} photo(s)" if photo_count else ""
            print(f"  Found {len(steps)} step(s){photo_info}")
    elif ext == ".xlsx":
        steps = extract_steps_from_excel(file_path)
        if steps:
            print(f"  Found {len(steps)} step(s)")

    if not steps:
        print("  No procedure steps found — skipping")
        return None, "No procedure steps found"

    # ── 2. Create IGT Standard record ─────────────────────────────────────────
    print("  Creating IGT Standard...", end=" ", flush=True)
    sys_id, number = create_igt_standard(instance, username, password, title, assign_type)
    if not sys_id:
        print(f"FAILED ({str(number)[:100]})")
        return None, str(number)[:200]
    print(f"OK ({number})")

    # ── 3. Create assessment questions — one Radio question per step ──────────
    print(f"  Creating {len(steps)} question(s)...", end=" ", flush=True)

    # The IGT Standard auto-creates a linked sn_smart_asmt_template record
    tmpl_sys_id = get_igt_assessment_template(instance, username, password, sys_id)
    if not tmpl_sys_id:
        print("SKIPPED (could not get assessment template)")
        return number, None

    # One section per unique ETAPE value — tracks {etape_title: section_sys_id}
    sections_by_etape = {}
    section_order     = 0

    _CT_EXT = {"image/jpeg": "jpg", "image/png": "png", "image/gif": "gif", "image/webp": "webp"}

    ok_q = fail_q = 0
    for step in steps:
        # ── Section: create once per unique ETAPE title ───────────────────────
        etape = step["title"] or "Général"
        if etape not in sections_by_etape:
            section_order += 1
            sec_id = create_igt_section(
                instance, username, password, tmpl_sys_id,
                name=etape, order=section_order,
            )
            if not sec_id:
                fail_q += 1
                continue
            sections_by_etape[etape] = sec_id
        section_sys_id = sections_by_etape[etape]

        # ── Question: METHODE instruction is the question text ────────────────
        # Fall back to ETAPE title or a numbered placeholder if instructions empty.
        label = step["instructions"] or step["title"] or f"Étape {step['order']}"

        q_sys_id, _ = create_igt_question(
            instance, username, password,
            template_sys_id=tmpl_sys_id,
            section_sys_id=section_sys_id,
            label=label,
            guidance_html="",   # photos uploaded separately below
            order=step["order"],
        )
        if not q_sys_id:
            fail_q += 1
            continue
        ok_q += 1

        # ── Response options: "Fait" (10) and "Non-fait" (20) ─────────────────
        create_igt_response_option(instance, username, password,
                                   q_sys_id, tmpl_sys_id, "Fait",     10)
        create_igt_response_option(instance, username, password,
                                   q_sys_id, tmpl_sys_id, "Non-fait", 20)

        # ── Guidance: upload PHOTO column images as attachments ───────────────
        if step.get("photos"):
            img_tags = []
            for i, (photo_bytes, ct) in enumerate(step["photos"], 1):
                ext_str = _CT_EXT.get(ct, "jpg")
                fname   = f"step_{step['order']}_photo{i}.{ext_str}"
                att_url = upload_attachment(
                    instance, username, password,
                    q_sys_id, fname, photo_bytes, ct,
                    table_name="sn_smart_asmt_question",
                )
                if att_url:
                    img_tags.append(
                        f'<img src="{att_url}" alt="{fname}" style="max-width:100%;" />'
                    )
            if img_tags:
                update_igt_question(
                    instance, username, password,
                    q_sys_id, "".join(img_tags),
                )

    status = f"{ok_q} OK"
    if fail_q:
        status += f", {fail_q} FAILED"
    print(status)
    print(f"  Sections created: {len(sections_by_etape)}")

    return number, None


def main():
    if len(sys.argv) < 2:
        print("Usage: python igt_to_kb.py <folder_path>")
        sys.exit(1)

    folder_path = Path(sys.argv[1])
    if not folder_path.is_dir():
        print(f"ERROR: '{folder_path}' is not a valid directory.")
        sys.exit(1)

    instance, username, password, assign_type = get_config()

    files = [
        f for f in folder_path.iterdir()
        if f.is_file() and f.suffix.lower() in SUPPORTED_EXTENSIONS
    ]

    if not files:
        print(f"No supported files found in '{folder_path}'.")
        print(f"Supported: {', '.join(sorted(SUPPORTED_EXTENSIONS))}")
        sys.exit(0)

    print(f"Found {len(files)} file(s) to process as IGT Standards.")
    print(f"Assignment type: {assign_type}\n")

    results = {"success": [], "failed": []}

    for file_path in sorted(files):
        print(f"[igt] {file_path.name}")
        try:
            number, error = process_igt(file_path, instance, username, password, assign_type)
            if number:
                results["success"].append((file_path.name, number))
            else:
                results["failed"].append((file_path.name, error))
        except Exception as e:
            print(f"  ERROR: {e}")
            results["failed"].append((file_path.name, str(e)))
        print()

    print("--- Summary ---")
    print(f"Successful: {len(results['success'])}")
    print(f"Failed:     {len(results['failed'])}")

    if results["success"]:
        print("\nCreated IGT Standards:")
        for name, number in results["success"]:
            print(f"  - {name} -> {number}")

    if results["failed"]:
        print("\nFailed files:")
        for name, reason in results["failed"]:
            print(f"  - {name}: {reason}")


if __name__ == "__main__":
    main()
